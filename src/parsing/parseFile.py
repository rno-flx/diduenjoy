# Function that parses the file and returns the data in variables

import openpyxl

def parseFile(file):
    dataframe = openpyxl.load_workbook(file)
    data = []
    
    for sheet in dataframe.sheetnames:
        worksheet = dataframe[sheet]
        for row in range(2, worksheet.max_row + 1):
            row_data = {
                'id': worksheet.cell(row=row, column=1).value,
                'name': worksheet.cell(row=row, column=2).value,
                'price': worksheet.cell(row=row, column=3).value,
                'ref': worksheet.cell(row=row, column=4).value,
                'packageId': worksheet.cell(row=row, column=5).value,
                'warranty': worksheet.cell(row=row, column=6).value,
                'duration': worksheet.cell(row=row, column=7).value
            }
            data.append(row_data)
    
    return data

